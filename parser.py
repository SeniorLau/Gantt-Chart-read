from openpyxl import load_workbook
import pandas as pd


def normalize_progress(value):

    if value is None:
        return "0%"

    try:

        # Numeric values
        if isinstance(value, (int, float)):

            # Excel percentage stored as decimal
            # 1 = 100%, 0.5 = 50%
            if 0 <= value <= 1:
                return f"{round(value * 100)}%"

            # Already percentage number
            return f"{round(value)}%"


        # Text values
        text = str(value).strip()

        text = text.replace("%", "")

        number = float(text)

        if 0 <= number <= 1:
            number = number * 100

        return f"{round(number)}%"


    except Exception:

        return "0%"



def parse_excel(filename):

    wb = load_workbook(
        filename,
        data_only=True
    )

    ws = wb.active
    wb_formula = load_workbook(
    filename,
    data_only=False
    )

    ws_formula = wb_formula.active


    tasks = []

    phase = None

    header_found = False


    for row_number in range(1, ws.max_row + 1):

        # Excel columns
        # B = Task
        # C = Assigned
        # D = Progress
        # E = Start
        # F = Finish

        task_cell = ws.cell(row_number, 2)
        assigned_cell = ws.cell(row_number, 3)
        progress_cell = ws.cell(row_number, 4)
        start_cell = ws.cell(row_number, 5)
        finish_cell = ws.cell(row_number, 6)


        b = task_cell.value
        c = assigned_cell.value
        d = progress_cell.value
        # Read progress from the original cell
        progress_raw = ws_formula.cell(
            row=row_number,
            column=4
        ).value
        
        
        # If formula exists, try the cached value first
        if isinstance(progress_raw, str) and progress_raw.startswith("="):
            d = progress_cell.value
        else:
            d = progress_raw
        e = start_cell.value
        f = finish_cell.value



        # Find table header

        if (
            str(b).strip() == "TAAK"
            and "TOEGEWEZEN" in str(c)
        ):

            header_found = True
            continue



        if not header_found:
            continue



        # Stop at bottom

        if (
            b
            and "Voeg nieuwe rijen" in str(b)
        ):
            break



        # Ignore empty rows

        if b is None:
            continue



        # Convert dates

        start = pd.to_datetime(
            e,
            dayfirst=True,
            errors="coerce"
        )

        finish = pd.to_datetime(
            f,
            dayfirst=True,
            errors="coerce"
        )



        # Phase row

        if (
            pd.isna(start)
            and pd.isna(finish)
            and b
        ):

            phase = str(b).strip()

            continue



        # Task row

        if (
            pd.notna(start)
            and pd.notna(finish)
        ):


            # Handle Excel percentage formatting

            progress_value = d

            if (
                "%" in str(progress_cell.number_format)
                and isinstance(progress_value, (int, float))
            ):

                progress_value = progress_value * 100



            tasks.append(
                {
                    "Phase": phase,

                    "Task": str(b).strip(),

                    "Assigned": (
                        str(c).strip()
                        if c
                        else "Unknown"
                    ),

                    "Progress_raw": d,

                    "Progress": normalize_progress(
                        progress_value
                    ),

                    "Start": start,

                    "Finish": finish,
                }
            )



    df = pd.DataFrame(tasks)


    return df
